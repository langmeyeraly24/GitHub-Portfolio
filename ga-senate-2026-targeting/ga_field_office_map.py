"""
GA Field Office Map  (v3 — side-by-side edition)
=================================================
Left panel:  Media markets (DMA fills) + field offices + major cities
Right panel: County partisan lean (muted palette) + field offices + major cities

Source data: GA County Master w_ Media Market (8).xlsx
Run from the CMI Campaign Plan directory.
"""

import os, glob, zipfile, urllib.request
import openpyxl
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.patheffects as pe
from matplotlib.lines import Line2D

# ── workbook ──────────────────────────────────────────────────────────────────
_candidates = sorted(glob.glob("GA County Master w_ Media Market*.xlsx"))
if not _candidates:
    raise FileNotFoundError("No 'GA County Master w_ Media Market*.xlsx' found. "
                            "Run from the CMI Campaign Plan directory.")
SRC_WB = _candidates[-1]
print(f"Using workbook: {SRC_WB}")

OUT_PNG             = "ga_field_office_map.png"
SHAPEFILE_CACHE_DIR = "shapefile_cache"

# ── muted partisan lean palette ───────────────────────────────────────────────
LEAN_COLORS = {
    "Strong D":    "#5b8db8",   # muted steel blue
    "Lean D":      "#a8c8e0",   # soft sky blue
    "Competitive": "#e8d87a",   # muted gold
    "Lean R":      "#e8a87a",   # muted peach-orange
    "Strong R":    "#c45c5c",   # muted brick red
    "Unknown":     "#d4d4d4",
}
LEAN_ORDER = ["Strong D", "Lean D", "Competitive", "Lean R", "Strong R"]

# ── TV DMA fills (left panel) — soft, distinct pastels ───────────────────────
DMA_FILL_COLORS = {
    "Atlanta":                  "#b8cce4",   # soft blue-grey
    "Chattanooga":              "#d5b8e4",   # soft lavender
    "Greenville-Spartanburg":   "#b8d5e4",   # pale teal-blue
    "Augusta-Aiken":            "#b8d4c8",   # pale sage
    "Macon":                    "#e4d8b8",   # pale straw
    "Columbus-Opelika":         "#e4c8b8",   # pale peach
    "Albany":                   "#c8e4b8",   # pale green
    "Savannah":                 "#b8e4d8",   # pale mint
    "Jacksonville":             "#d4d4d4",   # light grey
    "Tallahassee-Thomasville":  "#e4b8d8",   # pale rose
    "Dothan":                   "#e4e4b8",   # pale yellow-green
    "Unknown":                  "#ececec",
}

# DMA outline/label colors — darker version of the fill for contrast
DMA_EDGE_COLORS = {
    "Atlanta":                  "#3a5f82",
    "Chattanooga":              "#6a3a82",
    "Greenville-Spartanburg":   "#2a6070",
    "Augusta-Aiken":            "#2a5a48",
    "Macon":                    "#6a5a20",
    "Columbus-Opelika":         "#6a3820",
    "Albany":                   "#2a5a20",
    "Savannah":                 "#1a5848",
    "Jacksonville":             "#505060",
    "Tallahassee-Thomasville":  "#6a2a58",
    "Dothan":                   "#5a5a10",
}

# ── field offices ─────────────────────────────────────────────────────────────
FIELD_OFFICES = {
    "Athens HQ":       {"lat": 33.96,  "lon": -83.38},
    "North Atlanta":   {"lat": 34.07,  "lon": -84.51},
    "South Atlanta":   {"lat": 33.38,  "lon": -84.78},
    "Augusta":         {"lat": 33.47,  "lon": -82.01},
    "Savannah":        {"lat": 32.08,  "lon": -81.10},
    "Columbus/Albany": {"lat": 31.58,  "lon": -84.16},
    "Valdosta/S. GA":  {"lat": 30.83,  "lon": -83.28},
    "Macon":           {"lat": 32.84,  "lon": -83.63},
}

FO_COLORS = {
    "Athens HQ":       "#FFD700",
    "North Atlanta":   "#7b4faa",
    "South Atlanta":   "#5a3080",
    "Augusta":         "#2060a0",
    "Savannah":        "#0a8070",
    "Columbus/Albany": "#c05010",
    "Valdosta/S. GA":  "#a03008",
    "Macon":           "#1a7a3a",
}

# ── top-5 major cities ────────────────────────────────────────────────────────
MAJOR_CITIES = {
    "Atlanta":  {"lat": 33.749, "lon": -84.388},
    "Augusta":  {"lat": 33.474, "lon": -82.009},
    "Columbus": {"lat": 32.460, "lon": -84.987},
    "Savannah": {"lat": 32.080, "lon": -81.099},
    "Macon":    {"lat": 32.840, "lon": -83.632},
}

# ── TV DMA county assignments ─────────────────────────────────────────────────
TV_MARKETS = {
    "Atlanta": [
        "Banks","Barrow","Bartow","Butts","Carroll","Chattooga","Cherokee",
        "Clarke","Clayton","Cobb","Coweta","Dawson","DeKalb","Douglas",
        "Fannin","Fayette","Floyd","Forsyth","Fulton","Gilmer","Gordon",
        "Greene","Gwinnett","Habersham","Hall","Haralson","Heard","Henry",
        "Jackson","Jasper","Lamar","Lumpkin","Madison","Meriwether","Morgan",
        "Newton","Oconee","Oglethorpe","Paulding","Pickens","Pike","Polk",
        "Putnam","Rabun","Rockdale","Spalding","Towns","Troup","Union",
        "Upson","Walton","White",
    ],
    "Chattanooga":              ["Dade","Catoosa","Walker","Whitfield","Murray"],
    "Greenville-Spartanburg":   ["Elbert","Franklin","Hart","Stephens"],
    "Augusta-Aiken": [
        "Burke","Columbia","Emanuel","Glascock","Jefferson","Jenkins",
        "Lincoln","McDuffie","Richmond","Taliaferro","Warren","Wilkes",
    ],
    "Macon": [
        "Baldwin","Bibb","Bleckley","Crawford","Dodge","Dooly","Hancock",
        "Houston","Johnson","Jones","Laurens","Macon","Monroe","Montgomery",
        "Peach","Pulaski","Taylor","Telfair","Treutlen","Twiggs",
        "Washington","Wheeler","Wilcox","Wilkinson",
    ],
    "Columbus-Opelika": [
        "Chattahoochee","Clay","Harris","Marion","Muscogee","Quitman",
        "Randolph","Schley","Stewart","Sumter","Talbot","Webster",
    ],
    "Albany": [
        "Atkinson","Baker","Ben Hill","Berrien","Calhoun","Coffee",
        "Colquitt","Cook","Crisp","Dougherty","Irwin","Lanier","Lee",
        "Mitchell","Terrell","Tift","Turner","Worth",
    ],
    "Savannah": [
        "Appling","Bryan","Bulloch","Candler","Chatham","Effingham",
        "Evans","Jeff Davis","Liberty","Long","McIntosh","Screven",
        "Tattnall","Toombs","Wayne",
    ],
    "Jacksonville":             ["Bacon","Brantley","Camden","Charlton","Glynn","Pierce","Ware"],
    "Tallahassee-Thomasville":  ["Brooks","Clinch","Decatur","Echols","Grady","Lowndes","Seminole","Thomas"],
    "Dothan":                   ["Early","Miller"],
}

DMA_SHORT = {
    "Atlanta":                  "Atlanta DMA",
    "Chattanooga":              "Chattanooga",
    "Greenville-Spartanburg":   "Grv-Spbg",
    "Augusta-Aiken":            "Augusta-Aiken",
    "Macon":                    "Macon DMA",
    "Columbus-Opelika":         "Columbus-Op.",
    "Albany":                   "Albany DMA",
    "Savannah":                 "Savannah DMA",
    "Jacksonville":             "Jacksonville",
    "Tallahassee-Thomasville":  "Tallahassee",
    "Dothan":                   "Dothan",
}

# ── data loading ──────────────────────────────────────────────────────────────

def load_county_data():
    wb = openpyxl.load_workbook(SRC_WB, data_only=True)
    ws = wb['county_data_with_markets']
    rows = []
    for r in range(2, 161):
        county = ws.cell(row=r, column=1).value
        if not county:
            continue
        rows.append({
            "county":      county,
            "lean_bucket": ws.cell(row=r, column=51).value or "Unknown",
        })
    return rows


def load_ga_shapes():
    import geopandas as gpd
    os.makedirs(SHAPEFILE_CACHE_DIR, exist_ok=True)
    cache = os.path.join(SHAPEFILE_CACHE_DIR, "ga_counties.gpkg")
    if os.path.exists(cache):
        return gpd.read_file(cache)
    print("Downloading GA county shapefile (~5 MB, one-time)...")
    url = "https://www2.census.gov/geo/tiger/GENZ2023/shp/cb_2023_us_county_500k.zip"
    with urllib.request.urlopen(url) as resp:
        data = resp.read()
    zpath = os.path.join(SHAPEFILE_CACHE_DIR, "cb_us_county.zip")
    with open(zpath, "wb") as f:
        f.write(data)
    with zipfile.ZipFile(zpath) as z:
        z.extractall(SHAPEFILE_CACHE_DIR)
    shp = os.path.join(SHAPEFILE_CACHE_DIR, "cb_2023_us_county_500k.shp")
    counties = gpd.read_file(shp)
    ga = counties[counties["STATEFP"] == "13"].copy()
    ga["county"] = ga["NAME"]
    ga.to_file(cache, driver="GPKG")
    return ga


# ── shared overlay helpers ────────────────────────────────────────────────────

def draw_cities(ax):
    for city, info in MAJOR_CITIES.items():
        ax.scatter(info["lon"], info["lat"],
                   marker="o", s=55, c="#ffffff",
                   edgecolors="#222222", linewidths=1.6, zorder=14)
        ax.annotate(city, (info["lon"], info["lat"]),
                    xytext=(0, 6), textcoords="offset points",
                    fontsize=7.5, ha="center", fontweight="bold",
                    color="#111111",
                    path_effects=[pe.withStroke(linewidth=2.0, foreground="white")],
                    zorder=15)


def draw_field_offices(ax):
    for name, info in FIELD_OFFICES.items():
        color  = FO_COLORS[name]
        is_hq  = "HQ" in name
        marker = "*" if is_hq else "D"
        size   = 600 if is_hq else 160
        ax.scatter(info["lon"], info["lat"],
                   marker=marker, s=size, c=color,
                   edgecolors="#111111", linewidths=1.5, zorder=16)
        # label — Macon office sits on top of the city dot, push label right
        if name == "Macon":
            xoff, yoff, ha = 6, -12, "left"
        elif is_hq:
            xoff, yoff, ha = -7, -10, "right"
        else:
            xoff, yoff, ha = 6, 6, "left"
        ax.annotate(name, (info["lon"], info["lat"]),
                    xytext=(xoff, yoff), textcoords="offset points",
                    fontsize=7, fontweight="bold", ha=ha,
                    color="#111111",
                    bbox=dict(boxstyle="round,pad=0.25", facecolor="white",
                              edgecolor=color, linewidth=1.2, alpha=0.92),
                    zorder=17)


# ── main ──────────────────────────────────────────────────────────────────────

def make_map():
    import geopandas as gpd
    import pandas as pd

    print("Loading county data...")
    df = pd.DataFrame(load_county_data())

    print("Loading shapes...")
    ga = load_ga_shapes()
    ga = ga.merge(df, on="county", how="left")
    ga["lean_bucket"] = ga["lean_bucket"].fillna("Unknown")

    county_to_dma = {c: dma for dma, cs in TV_MARKETS.items() for c in cs}
    ga["tv_market"] = ga["county"].map(county_to_dma).fillna("Unknown")

    dma_shapes = ga.dissolve(by="tv_market").reset_index()

    # ── figure: two panels side by side ──────────────────────────────────────
    fig, (ax_left, ax_right) = plt.subplots(1, 2, figsize=(22, 12))
    fig.patch.set_facecolor("white")

    for ax in (ax_left, ax_right):
        ax.set_facecolor("#eef2f5")
        ax.set_axis_off()

    # ── LEFT PANEL: media markets ─────────────────────────────────────────────

    # Fill by DMA
    for dma, fill in DMA_FILL_COLORS.items():
        subset = ga[ga["tv_market"] == dma]
        if len(subset):
            subset.plot(ax=ax_left, color=fill, edgecolor="#bbbbbb",
                        linewidth=0.3, alpha=0.95)

    # DMA boundary outlines (thick)
    for _, row in dma_shapes.iterrows():
        dma = row["tv_market"]
        if dma == "Unknown":
            continue
        edge = DMA_EDGE_COLORS.get(dma, "#444444")
        gpd.GeoSeries([row.geometry]).plot(
            ax=ax_left, facecolor="none", edgecolor=edge,
            linewidth=2.0, zorder=5
        )

    # DMA name labels
    for _, row in dma_shapes.iterrows():
        dma = row["tv_market"]
        if dma == "Unknown":
            continue
        centroid = row.geometry.centroid
        edge = DMA_EDGE_COLORS.get(dma, "#333333")
        ax_left.text(centroid.x, centroid.y, DMA_SHORT.get(dma, dma),
                     fontsize=6.5, ha="center", va="center",
                     color=edge, fontweight="bold", alpha=0.85,
                     path_effects=[pe.withStroke(linewidth=1.8, foreground="white")],
                     zorder=6)

    draw_cities(ax_left)
    draw_field_offices(ax_left)

    ax_left.set_title("TV Media Markets & Field Offices",
                      fontsize=12, fontweight="bold", pad=10, color="#1a1a1a")

    # DMA legend (bottom-left)
    dma_handles = [
        mpatches.Patch(facecolor=DMA_FILL_COLORS[k],
                       edgecolor=DMA_EDGE_COLORS.get(k, "#444"),
                       linewidth=1.2,
                       label=DMA_SHORT[k])
        for k in DMA_EDGE_COLORS
    ]
    ax_left.legend(handles=dma_handles, loc="lower left", fontsize=7,
                   title="Media Market (DMA)", title_fontsize=8,
                   frameon=True, framealpha=0.95, edgecolor="#bbb",
                   ncol=1)

    # ── RIGHT PANEL: partisan lean ────────────────────────────────────────────

    for bucket in LEAN_ORDER:
        subset = ga[ga["lean_bucket"] == bucket]
        if len(subset):
            subset.plot(ax=ax_right, color=LEAN_COLORS[bucket],
                        edgecolor="#aaaaaa", linewidth=0.3, alpha=0.90)
    unknown = ga[ga["lean_bucket"] == "Unknown"]
    if len(unknown):
        unknown.plot(ax=ax_right, color=LEAN_COLORS["Unknown"],
                     edgecolor="#aaaaaa", linewidth=0.3, alpha=0.90)

    draw_cities(ax_right)
    draw_field_offices(ax_right)

    ax_right.set_title("County Partisan Lean & Field Offices",
                       fontsize=12, fontweight="bold", pad=10, color="#1a1a1a")

    # Lean legend (bottom-left of right panel)
    lean_handles = [
        mpatches.Patch(facecolor=LEAN_COLORS[b], edgecolor="#888",
                       linewidth=0.8, label=b)
        for b in LEAN_ORDER
    ]
    ax_right.legend(handles=lean_handles, loc="lower left", fontsize=8,
                    title="Partisan Lean", title_fontsize=9,
                    frameon=True, framealpha=0.95, edgecolor="#bbb")

    # ── shared field-office + city legend (bottom of right panel) ────────────
    fo_handles = []
    for name, color in FO_COLORS.items():
        is_hq  = "HQ" in name
        marker = "*" if is_hq else "D"
        fo_handles.append(
            plt.scatter([], [], marker=marker,
                        s=100 if is_hq else 50,
                        c=color, edgecolors="#111", linewidths=1.0,
                        label=name)
        )
    fo_handles.append(
        plt.scatter([], [], marker="o", s=45,
                    c="#ffffff", edgecolors="#222222", linewidths=1.4,
                    label="Top-5 city")
    )
    ax_right.legend(handles=fo_handles, loc="lower right", fontsize=7.5,
                    title="Field Offices & Cities", title_fontsize=8.5,
                    frameon=True, framealpha=0.95, edgecolor="#bbb")

    # Re-add lean legend (matplotlib replaces on second call)
    lean_leg = mpatches.Patch  # just a ref; rebuild below
    lean_handles2 = [
        mpatches.Patch(facecolor=LEAN_COLORS[b], edgecolor="#888",
                       linewidth=0.8, label=b)
        for b in LEAN_ORDER
    ]
    leg_lean = ax_right.legend(handles=lean_handles2, loc="lower left",
                               fontsize=8, title="Partisan Lean",
                               title_fontsize=9, frameon=True,
                               framealpha=0.95, edgecolor="#bbb")
    # Then add FO legend without clobbering lean
    from matplotlib.legend import Legend
    leg_fo = Legend(ax_right, fo_handles,
                    [h.get_label() for h in fo_handles],
                    loc="lower right", fontsize=7.5,
                    title="Field Offices & Cities", title_fontsize=8.5,
                    frameon=True, framealpha=0.95, edgecolor="#bbb")
    ax_right.add_artist(leg_lean)
    ax_right.add_artist(leg_fo)

    # ── overall title ─────────────────────────────────────────────────────────
    fig.suptitle("Georgia 2026 Senate — Collins Campaign Field Plan",
                 fontsize=15, fontweight="bold", y=0.98, color="#1a1a1a")

    plt.tight_layout(rect=[0, 0, 1, 0.96])
    plt.savefig(OUT_PNG, dpi=200, bbox_inches="tight", facecolor="white")
    plt.close()
    print(f"Saved: {OUT_PNG}")


if __name__ == "__main__":
    make_map()
