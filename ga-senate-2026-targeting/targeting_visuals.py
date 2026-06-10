"""
Targeting Section - Visualization Generator
============================================
Generates all 5 PNG charts used in Targeting_v3.docx.

HOW TO USE THIS FILE IN POSITRON:
  1. Open in Positron, run the whole file (Cmd/Ctrl+Shift+Enter).
  2. The 5 viz_*.png files write out to the same directory as this script.
  3. To tweak a single chart, find its function below and edit, then re-run just
     that function (or rerun the whole file).
  4. After regenerating a PNG, replace the embedded image in Targeting_v3.docx
     by opening the doc in Word, right-clicking the chart, and "Change Picture".

QUICK-EDIT CONFIG: see the CONFIG block immediately below this docstring.
Everything you'd typically want to change (fig sizes, colors, fonts) is there.

Data source:  GA County Master w_ Media Market (8).xlsx
              (Summary + county_data_with_markets tabs)
Requirements: openpyxl, matplotlib, numpy
Author:       Alyssa Langmeyer (Collins for U.S. Senate 2026 targeting analysis)
"""

import openpyxl
import matplotlib
matplotlib.use('Agg')  # non-interactive backend; remove if you want windows in Positron
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np
from matplotlib.patches import Rectangle


# =============================================================================
# CONFIG  --  EDIT HERE FOR QUICK CHANGES
# =============================================================================

# Where the source workbook lives (relative to this script)
SRC_WB = "GA County Master w_ Media Market (8).xlsx"

# Where PNGs are saved
OUT_DIR = "."

# Resolution (DPI) for all PNG output. 170 = good for embedding in Word.
DPI = 170

# Figure sizes (width, height) in inches. Change these to make charts bigger/smaller.
FIG_SIZE_MATRIX        = (15, 8.5)   # the 5x3 targeting matrix
FIG_SIZE_TOP10_PRIO    = (11, 6.5)
FIG_SIZE_KW_GAP        = (13, 6.5)
FIG_SIZE_PATH          = (11, 5.5)
FIG_SIZE_LEAN_BUCKETS  = (13, 5.5)

# Targeting matrix cell size — wider cells fit the long "MOBILIZE + LIGHT PERSUADE"
# label better. Change CELL_W to make cells narrower/wider.
MATRIX_CELL_W = 1.6
MATRIX_CELL_H = 1.0

# Color palette — partisan red + neutral
R_RED       = "#C0392B"
R_RED_LIGHT = "#E74C3C"
NEUTRAL     = "#95A5A6"
ORANGE      = "#E67E22"

# Action colors (used in matrix). Edit hex codes to recolor any action.
action_colors = {
    "MAINTAIN":                  "#27AE60",  # green
    "MOBILIZE":                  "#C0392B",  # red
    "MOBILIZE + LIGHT PERSUADE": "#E67E22",  # orange
    "PERSUADE":                  "#F39C12",  # yellow
    "DEFENSE":                   "#3498DB",  # blue
    "DEFENSE / MIN INVEST":      "#5DADE2",  # light blue
    "SKIP":                      "#BDC3C7",  # gray
}

# Lean bucket colors (used in lean overview + KW gap)
lean_color_map = {
    "Strong R":    "#922B21",
    "Lean R":      "#C0392B",
    "Competitive": "#F39C12",
    "Lean D":      "#2874A6",
    "Strong D":    "#1B4F72",
}

# Wrap long action labels onto two lines so they fit in matrix cells
def wrap_action(s):
    if s == "MOBILIZE + LIGHT PERSUADE": return "MOBILIZE +\nLIGHT PERSUADE"
    if s == "DEFENSE / MIN INVEST":      return "DEFENSE /\nMIN INVEST"
    return s

# Global matplotlib style
plt.rcParams['font.family'] = 'DejaVu Sans'
plt.rcParams['axes.spines.top'] = False
plt.rcParams['axes.spines.right'] = False


# =============================================================================
# DATA EXTRACTION (you typically don't need to edit this)
# =============================================================================

def extract_facts(src_path):
    """Pull every number used in the visualizations from the workbook."""
    wb = openpyxl.load_workbook(src_path, data_only=True)
    ws_sum = wb['Summary']
    ws_cdm = wb['county_data_with_markets']

    F = {
        "total_counties": int(ws_sum['B1'].value),
        "total_expected_voters": int(ws_sum['B4'].value),
        "win_number_50plus1": int(ws_sum['B5'].value),
        "target_52pct": round(ws_sum['B6'].value),
        "baseline_collins_walker": round(ws_sum['B7'].value),
        "persuasion_universe_total": int(ws_sum['B9'].value),
        "total_active_voters": 0,
    }

    # Lean bucket aggregates
    F["lean_buckets"] = {}
    for r in range(2, 161):
        bucket = ws_cdm.cell(row=r, column=51).value  # lean_bucket
        if bucket not in F["lean_buckets"]:
            F["lean_buckets"][bucket] = {"counties": 0, "active_voters": 0,
                                          "persuasion": 0, "mobilization": 0}
        F["lean_buckets"][bucket]["counties"] += 1
        F["lean_buckets"][bucket]["active_voters"] += ws_cdm.cell(row=r, column=24).value or 0
        F["lean_buckets"][bucket]["persuasion"] += ws_cdm.cell(row=r, column=54).value or 0
        F["lean_buckets"][bucket]["mobilization"] += ws_cdm.cell(row=r, column=55).value or 0
    F["total_active_voters"] = sum(b["active_voters"] for b in F["lean_buckets"].values())

    # 5x3 targeting matrix (lean x turnout)
    F["matrix"] = {}
    for r in range(2, 161):
        bucket = ws_cdm.cell(row=r, column=51).value
        tier = ws_cdm.cell(row=r, column=52).value
        key = f"{bucket} | {tier}"
        if key not in F["matrix"]:
            F["matrix"][key] = {"counties": 0, "active": 0, "mobilization": 0, "persuasion": 0,
                                "action": ws_cdm.cell(row=r, column=53).value}
        F["matrix"][key]["counties"] += 1
        F["matrix"][key]["active"] += ws_cdm.cell(row=r, column=24).value or 0
        F["matrix"][key]["mobilization"] += ws_cdm.cell(row=r, column=55).value or 0
        F["matrix"][key]["persuasion"] += ws_cdm.cell(row=r, column=54).value or 0

    # Top 10 Kemp-Walker gap (col 49 = kemp_walker_gap_pts)
    kw = []
    for r in range(2, 161):
        kw.append({"county": ws_cdm.cell(row=r, column=1).value,
                   "gap_pts": ws_cdm.cell(row=r, column=49).value,
                   "lean": ws_cdm.cell(row=r, column=51).value,
                   "active": ws_cdm.cell(row=r, column=24).value,
                   "persuasion": ws_cdm.cell(row=r, column=54).value})
    F["top10_kw"] = sorted([k for k in kw if k["gap_pts"] is not None],
                            key=lambda x: -x["gap_pts"])[:10]

    # Top 10 priority counties (col 57 = priority_rank)
    pr = []
    for r in range(2, 161):
        pr.append({"county": ws_cdm.cell(row=r, column=1).value,
                   "rank": ws_cdm.cell(row=r, column=57).value,
                   "action": ws_cdm.cell(row=r, column=53).value,
                   "persuasion": ws_cdm.cell(row=r, column=54).value or 0,
                   "mobilization": ws_cdm.cell(row=r, column=55).value or 0})
    F["top10_priority"] = sorted([p for p in pr if p["rank"] is not None],
                                  key=lambda x: x["rank"])[:10]
    return F


# =============================================================================
# CHART 1: TARGETING MATRIX (5x3 grid, color-coded by action)
# =============================================================================
# This is the one that was "squished on the x-axis" in the earlier version.
# Fix: wider cells (CELL_W in CONFIG) and removed set_aspect('equal').
# Edit MATRIX_CELL_W in CONFIG to change cell width.
# =============================================================================

def chart_targeting_matrix(F, out):
    lean_order = ["Strong R", "Lean R", "Competitive", "Lean D", "Strong D"]
    tier_order = ["Low", "Mid", "High"]

    fig, ax = plt.subplots(figsize=FIG_SIZE_MATRIX)
    CW, CH = MATRIX_CELL_W, MATRIX_CELL_H

    for i, lean in enumerate(lean_order):
        for j, tier in enumerate(tier_order):
            cell = F["matrix"].get(f"{lean} | {tier}",
                                    {"counties": 0, "active": 0, "mobilization": 0, "action": "-"})
            action = cell["action"] or "-"
            color = action_colors.get(action, "#FFFFFF")
            x = j * CW
            y = (len(lean_order)-1-i) * CH
            rect = Rectangle((x, y), CW, CH,
                             facecolor=color, edgecolor='white', linewidth=2)
            ax.add_patch(rect)
            # Action label (top)
            ax.text(x + CW/2, y + 0.78, wrap_action(action),
                    ha='center', va='center', fontsize=10, color='white', fontweight='bold')
            # Counties
            ax.text(x + CW/2, y + 0.50, f"{cell['counties']} counties",
                    ha='center', va='center', fontsize=11, color='white', fontweight='bold')
            # Active voters
            ax.text(x + CW/2, y + 0.30, f"{cell['active']:,} active",
                    ha='center', va='center', fontsize=9.5, color='white')
            # Mobilization universe
            ax.text(x + CW/2, y + 0.13, f"{cell['mobilization']:,} mob.",
                    ha='center', va='center', fontsize=9, color='white', style='italic')

    ax.set_xlim(-0.05, 3*CW + 0.05)
    ax.set_ylim(-0.05, 5*CH + 0.05)
    ax.set_xticks([CW*0.5, CW*1.5, CW*2.5])
    ax.set_xticklabels(["LOW turnout", "MID turnout", "HIGH turnout"],
                       fontsize=12, fontweight='bold')
    ax.set_yticks([0.5, 1.5, 2.5, 3.5, 4.5])
    ax.set_yticklabels(list(reversed(lean_order)), fontsize=12, fontweight='bold')
    # DELIBERATELY no set_aspect('equal') — that's what was squishing the old chart
    ax.set_title("Targeting Matrix: Lean Bucket x Turnout Tier\n"
                 "159 GA counties, color-coded by recommended action",
                 fontsize=14, fontweight='bold', pad=20)
    ax.tick_params(axis='both', length=0)
    ax.spines[:].set_visible(False)

    # Legend
    legend_actions = list(action_colors.keys())
    patches = [Rectangle((0,0),1,1, facecolor=action_colors[a]) for a in legend_actions]
    ax.legend(patches, legend_actions, loc='center left', bbox_to_anchor=(1.02, 0.5),
              frameon=False, fontsize=10, title="Recommended Action", title_fontsize=11)

    plt.tight_layout()
    plt.savefig(out, dpi=DPI, bbox_inches='tight', facecolor='white')
    plt.close()


# =============================================================================
# CHART 2: TOP 10 PRIORITY COUNTIES (horizontal stacked bars)
# =============================================================================

def chart_top10_priority(F, out):
    top10 = F["top10_priority"]
    counties = [c["county"] for c in top10]
    mob = [c["mobilization"] for c in top10]
    pers = [c["persuasion"] for c in top10]
    actions = [c["action"] for c in top10]

    fig, ax = plt.subplots(figsize=FIG_SIZE_TOP10_PRIO)
    y = np.arange(len(counties))[::-1]
    ax.barh(y, mob, color=R_RED, label="Mobilization universe (un-turned-out R voters)")
    ax.barh(y, pers, left=mob, color=ORANGE, label="Persuasion universe (Kemp-but-not-Walker pool)")
    ax.set_yticks(y); ax.set_yticklabels(counties, fontsize=10)
    ax.set_xlabel("Voters", fontsize=10)
    ax.set_title("Top 10 Priority Counties for Collins\nCombined mobilization + persuasion opportunity",
                 fontsize=12, fontweight='bold', pad=15)
    ax.legend(loc='lower right', fontsize=9, frameon=False)
    for i, (m, p, a) in enumerate(zip(mob, pers, actions)):
        ax.text(m + p + max(mob)*0.01, y[i], f"  {a}", va='center', fontsize=8,
                color='#444', style='italic')
    ax.set_xlim(0, max([m+p for m, p in zip(mob, pers)]) * 1.35)
    ax.grid(axis='x', alpha=0.3); ax.spines['left'].set_visible(False)
    ax.tick_params(axis='y', length=0)
    plt.tight_layout()
    plt.savefig(out, dpi=DPI, bbox_inches='tight', facecolor='white')
    plt.close()


# =============================================================================
# CHART 3: KEMP-WALKER GAP (two-panel — gap pts and persuasion universe)
# =============================================================================

def chart_kemp_walker(F, out):
    top10kw = F["top10_kw"]
    counties = [c["county"] for c in top10kw]
    gaps = [c["gap_pts"] for c in top10kw]
    pers = [c["persuasion"] for c in top10kw]
    colors = [lean_color_map.get(c["lean"], "#888") for c in top10kw]

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=FIG_SIZE_KW_GAP, gridspec_kw={'width_ratios': [1, 1.2]})
    y = np.arange(len(counties))[::-1]
    ax1.barh(y, gaps, color=colors)
    ax1.set_yticks(y); ax1.set_yticklabels(counties, fontsize=10)
    ax1.set_xlabel("Kemp R% minus Walker R% (percentage points)", fontsize=9.5)
    ax1.set_title("Kemp-Walker Gap (pts)", fontsize=11, fontweight='bold')
    ax1.grid(axis='x', alpha=0.3); ax1.spines['left'].set_visible(False)
    ax1.tick_params(axis='y', length=0)
    for i, g in enumerate(gaps):
        ax1.text(g + 0.1, y[i], f"{g}", va='center', fontsize=8.5)
    ax2.barh(y, pers, color=colors)
    ax2.set_yticks(y); ax2.set_yticklabels([])
    ax2.set_xlabel("Estimated split-ticket voters (Kemp R, not Walker R)", fontsize=9.5)
    ax2.set_title("Persuasion Universe (voters)", fontsize=11, fontweight='bold')
    ax2.grid(axis='x', alpha=0.3); ax2.spines['left'].set_visible(False)
    ax2.tick_params(axis='y', length=0)
    for i, p in enumerate(pers):
        ax2.text(p + max(pers)*0.01, y[i], f"{p:,}", va='center', fontsize=8.5)
    handles = [mpatches.Patch(color=lean_color_map[k], label=k)
               for k in ["Strong R", "Lean R", "Competitive", "Lean D", "Strong D"]]
    fig.legend(handles=handles, loc='lower center', ncol=5, frameon=False,
               fontsize=9, bbox_to_anchor=(0.5, -0.02))
    fig.suptitle("Top 10 Counties by Kemp-Walker Gap: The Persuasion Target",
                 fontsize=13, fontweight='bold', y=1.00)
    plt.tight_layout()
    plt.savefig(out, dpi=DPI, bbox_inches='tight', facecolor='white')
    plt.close()


# =============================================================================
# CHART 4: PATH TO VICTORY (baseline vs win floor vs 52% target)
# =============================================================================

def chart_path_to_victory(F, out):
    fig, ax = plt.subplots(figsize=FIG_SIZE_PATH)
    cats = ["Walker Baseline\n(if Collins matches\nWalker's 2022 share)",
            "Win Floor\n(50% + 1, avoid runoff)",
            "Safety Target\n(52%, comfortable margin)"]
    vals = [F["baseline_collins_walker"], F["win_number_50plus1"], F["target_52pct"]]
    colors = [NEUTRAL, R_RED, R_RED_LIGHT]
    bars = ax.bar(cats, vals, color=colors, edgecolor='white', linewidth=2, width=0.6)
    for bar, val in zip(bars, vals):
        ax.text(bar.get_x() + bar.get_width()/2, val + 30000, f"{val:,}",
                ha='center', va='bottom', fontsize=11, fontweight='bold')
    g1 = vals[1] - vals[0]; g2 = vals[2] - vals[1]
    ax.annotate('', xy=(1, vals[1]), xytext=(0, vals[0]),
                arrowprops=dict(arrowstyle='->', color=R_RED, lw=2))
    ax.text(0.5, (vals[0]+vals[1])/2, f"+{g1:,}\nto avoid runoff",
            ha='center', fontsize=10, color=R_RED, fontweight='bold')
    ax.annotate('', xy=(2, vals[2]), xytext=(1, vals[1]),
                arrowprops=dict(arrowstyle='->', color=R_RED_LIGHT, lw=2))
    ax.text(1.5, (vals[1]+vals[2])/2, f"+{g2:,}\nbuffer to 52%",
            ha='center', fontsize=10, color=R_RED_LIGHT, fontweight='bold')
    ax.set_ylabel("Republican votes", fontsize=11)
    ax.set_title(f"Path to Victory: Where Collins Starts vs. Where He Needs to Be\n"
                 f"Total expected GA turnout: {F['total_expected_voters']:,}",
                 fontsize=12, fontweight='bold', pad=15)
    ax.set_ylim(0, vals[2] * 1.13)
    ax.grid(axis='y', alpha=0.3)
    ax.yaxis.set_major_formatter(plt.matplotlib.ticker.FuncFormatter(lambda x, _: f"{int(x/1000):,}k"))
    plt.tight_layout()
    plt.savefig(out, dpi=DPI, bbox_inches='tight', facecolor='white')
    plt.close()


# =============================================================================
# CHART 5: LEAN BUCKET OVERVIEW (counties vs voters + opportunity per bucket)
# =============================================================================

def chart_lean_buckets(F, out):
    lean_order = ["Strong R", "Lean R", "Competitive", "Lean D", "Strong D"]
    counties_per = [F["lean_buckets"][b]["counties"] for b in lean_order]
    active_per = [F["lean_buckets"][b]["active_voters"] for b in lean_order]
    pers_per = [F["lean_buckets"][b]["persuasion"] for b in lean_order]
    mob_per = [F["lean_buckets"][b]["mobilization"] for b in lean_order]
    total_active = sum(active_per)

    fig, axes = plt.subplots(1, 2, figsize=FIG_SIZE_LEAN_BUCKETS)
    x = np.arange(len(lean_order)); w = 0.38
    pct_c = [c/sum(counties_per)*100 for c in counties_per]
    pct_a = [a/total_active*100 for a in active_per]
    ax = axes[0]
    ax.bar(x - w/2, pct_c, w, label="% of counties (159 total)",
           color=[lean_color_map[b] for b in lean_order], alpha=0.6)
    ax.bar(x + w/2, pct_a, w, label="% of active voters (7.36M total)",
           color=[lean_color_map[b] for b in lean_order])
    for i, pct in enumerate(pct_c): ax.text(i - w/2, pct + 0.5, f"{pct:.0f}%", ha='center', fontsize=8.5)
    for i, pct in enumerate(pct_a): ax.text(i + w/2, pct + 0.5, f"{pct:.0f}%", ha='center', fontsize=8.5, fontweight='bold')
    ax.set_xticks(x); ax.set_xticklabels(lean_order, fontsize=10)
    ax.set_ylabel("% of total", fontsize=10)
    ax.set_title("Where Republicans win COUNTIES vs. where the VOTERS live", fontsize=11, fontweight='bold')
    ax.legend(loc='upper right', frameon=False, fontsize=9); ax.grid(axis='y', alpha=0.3)
    ax.set_ylim(0, max(max(pct_c), max(pct_a)) * 1.18)

    ax2 = axes[1]
    ax2.bar(x, pers_per, width=0.55, color="#E67E22", label="Persuasion universe")
    ax2.bar(x, mob_per, bottom=pers_per, width=0.55, color="#C0392B", label="Mobilization universe")
    for i, b in enumerate(lean_order):
        total = pers_per[i] + mob_per[i]
        ax2.text(i, total + 30000, f"{total:,}", ha='center', fontsize=9, fontweight='bold')
    ax2.set_xticks(x); ax2.set_xticklabels(lean_order, fontsize=10)
    ax2.set_ylabel("Voters", fontsize=10)
    ax2.set_title("Total opportunity (mobilization + persuasion) per bucket", fontsize=11, fontweight='bold')
    ax2.legend(loc='upper right', frameon=False, fontsize=9); ax2.grid(axis='y', alpha=0.3)
    ax2.yaxis.set_major_formatter(plt.matplotlib.ticker.FuncFormatter(lambda x, _: f"{int(x/1000):,}k"))
    fig.suptitle("Lean Bucket Overview: The Geography Problem and the Opportunity",
                 fontsize=13, fontweight='bold', y=1.02)
    plt.tight_layout()
    plt.savefig(out, dpi=DPI, bbox_inches='tight', facecolor='white')
    plt.close()


# =============================================================================
# MAIN — runs all 5 charts. Comment out any line to skip a chart.
# =============================================================================

def main():
    print(f"Reading {SRC_WB}...")
    F = extract_facts(SRC_WB)
    print(f"  Loaded {F['total_counties']} counties. Statewide active voters: {F['total_active_voters']:,}")
    print("Generating charts...")
    chart_targeting_matrix(F, f"{OUT_DIR}/viz_targeting_matrix.png")
    chart_top10_priority(F,    f"{OUT_DIR}/viz_top10_priority.png")
    chart_kemp_walker(F,       f"{OUT_DIR}/viz_kemp_walker_gap.png")
    chart_path_to_victory(F,   f"{OUT_DIR}/viz_path_to_victory.png")
    chart_lean_buckets(F,      f"{OUT_DIR}/viz_lean_buckets.png")
    print("Done. All 5 PNGs saved.")


if __name__ == "__main__":
    main()
